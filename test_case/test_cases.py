from key_word.webui_keyword import WebUIKeys, load_workbook
import openpyxl
from openpyxl.styles import PatternFill,Font
from log.log import Logger
log=Logger().log()
from common.common import ReadExcel2



class ExcelExcute:
    ec = ReadExcel2()

    def excute_sheets(self,excel_path):
        global excel
        #excel_path=excel_path
        excel = self.ec.load_excel(excel_path)
        sheets = self.ec.load_sheets(excel)
        return sheets





    def excute(self,excel_path):
        #workbook = openpyxl.load_workbook("../config/shopxo.xlsx")
        sheets = self.excute_sheets(excel_path)
        for sheet in sheets:
            sheet = excel[sheet]



        #for sheet_name in workbook.sheetnames:
            #sheet = workbook[sheet_name]
            for value in sheet.values:
                if type(value[1]) is int:
                    param_dict = dict.fromkeys(("type", "value", "text", "expect"))
                    param_dict["type"] = value[3]
                    param_dict["value"]=value[4]
                    param_dict["text"]=value[5]
                    param_dict["expect"]=value[7]
                    if value[2] == 'open_browser':
                        log.info('正在执行关键字:{0}，操作描述：{1}'.format(value[2], value[6]))
                        wk = WebUIKeys(value[5])
                    elif 'assert' in value[2]:
                        log.info('正在执行关键字:{0}，操作描述：{1}'.format(value[2], value[6]))
                        status = getattr(wk,value[2])(**param_dict)
                        if status is True:
                            row = value[1]+1
                            sheet.cell(row, 9).value = "pass"
                            sheet.cell(row, 9).fill = PatternFill('solid',fgColor='66ff00')
                            sheet.cell(row, 9).font = Font(bold=True)
                        else:
                            row = value[1] + 1
                            sheet.cell(row, 9).value = "false"
                            sheet.cell(row, 9).fill = PatternFill('solid', fgColor='FF0000')
                            sheet.cell(row, 9).font = Font(bold=True)

                    elif value[2] == 'quit':
                        log.info('正在执行关键字:{0}，操作描述：{1}'.format(value[2], value[6]))
                        getattr(wk, value[2])()
                    else:
                        log.info('正在执行关键字:{0}，操作描述：{1}'.format(value[2], value[6]))
                        getattr(wk, value[2])(**param_dict)
        self.ec.save_excel(excel, excel_path)
        log.info("文件读取完毕，自动化执行结束！\n")
        self.ec.close_excel(excel)


if __name__ == '__main__':
    ExcelExcute().excute("../config/shopxo.xlsx")
#v2.0
# workbook = load_workbook("../config/shopxo.xlsx")
# sheet = workbook["Sheet1"]
#
# for value in sheet.values:
#     if type(value[0]) is int:
#         if value[1] == 'open_browser':
#             wk = WebUIKeys(value[4])
#         elif 'assert' in value[1]:
#             log.info("调用assert_text关键字，描述信息：{}".format(value[5]))
#             status = getattr(wk,value[1])(value[2],value[3],value[6])
#             if status is True:
#                 row = value[0]+1#value[0]为当前关键字所在行的用例编号
#                 sheet.cell(row, 8).value="PASS"
#                 sheet.cell(row, 8).fill = PatternFill('solid',fgColor='66ff00')
#                 sheet.cell(row, 8).font = Font(bold=True)
#             elif status is False:
#                 row = value[0] + 1
#                 sheet.cell(row, 8).value = "FALSE"
#                 sheet.cell(row, 8).fill = PatternFill('solid', fgColor='cc0000')
#                 sheet.cell(row, 8).font = Font(bold=True)
#             workbook.save("../config/shopxo.xlsx")
#         else:
#             #getattr(wk,value[1])(value[2],value[3],value[4])
#             func = getattr(wk,value[1])
#             sum = func.__code__.co_argcount
#
#             if sum == 1:
#                 getattr(wk,value[1])()
#                 log.info("调用{1}关键字，所传参数个数为：{0}".format(sum,value[1]))
#             elif sum == 2:
#                 getattr(wk,value[1])(value[4])
#                 log.info("调用{1}关键字，所传参数个数为：{0}".format(sum, value[1]))
#             elif sum == 3:
#                 getattr(wk,value[1])(value[2],value[3])
#                 log.info("调用{1}关键字，所传参数个数为：{0}".format(sum, value[1]))
#             elif sum == 4:
#                 getattr(wk,value[1])(value[2],value[3],value[4])
#                 log.info("调用{1}关键字，所传参数个数为：{0}".format(sum, value[1]))
#             elif sum == 5:
#                 log.info("参数过多。参数个数为{}".format(sum))
#
#     else:
#         pass
#v1.0
# for value in sheet.values:#遍历得到一条条元组数据
#     if value[1] == 'open_browser':
#         log.info("实例化关键字对象，描述信息：{}".format(value[5]))
#         wk = WebUIKeys(value[4])
#
#     elif value[1] == "visit":
#         log.info("调用visit关键字，描述信息：{}".format(value[5]))
#         wk.visit(value[4])
#     elif value[1] == 'input':
#         log.info("调用visit关键字，描述信息：{}".format(value[5]))
#         wk.input(value[2],value[3],value[4])
#     elif value[1] == 'click':
#         log.info("调用click关键字，描述信息：{}".format(value[5]))
#         wk.click(value[2],value[3])
#     elif value[1] == 'wait':
#         log.info("调用wait关键字，描述信息：{}".format(value[5]))
#         wk.wait(value[4])
#     elif value[1] == 'assert_text':
#         log.info("调用assert_text关键字，描述信息：{}".format(value[5]))
#         status = wk.assert_text(value[2],value[3],value[6])
#         if status is True:
#             row = value[0]+1#value[0]为当前关键字所在行的用例编号
#             sheet.cell(row, 8).value="PASS"
#             sheet.cell(row, 8).fill = PatternFill('solid',fgColor='66ff00')
#             sheet.cell(row, 8).font = Font(bold=True)
#         elif status is False:
#             row = value[0] + 1
#             sheet.cell(row, 8).value = "FALSE"
#             sheet.cell(row, 8).fill = PatternFill('solid', fgColor='cc0000')
#             sheet.cell(row, 8).font = Font(bold=True)
#         workbook.save("../config/shopxo.xlsx")
#
#     else:
#         pass

